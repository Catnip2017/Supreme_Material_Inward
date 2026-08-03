"""
robot_scripts/dms_excel_writer.py — appends one {File Name, Document Link}
row to the DMS links Excel file, for dms_upload.robot's
"Generate And Save Document Link" step.

Config-driven replacement for the original dms_bot/excel_writer.py (which
hardcoded EXCEL_PATH to a personal C:\\Users\\ctn_ravi\\Downloads\\...
path). The path is passed in as a command-line argument
(config.DMS_LINKS_EXCEL_PATH, read via .env by the robot -- see
DMS_LINKS_EXCEL_PATH in .env.template) so it never drifts out of sync with
the app's actual configured location. Falls back to the app's default path
only if called without the third argument.

Usage: dms_excel_writer.py <file_name> <doc_link> [excel_path]
"""

import sys
import os
from openpyxl import Workbook, load_workbook

DEFAULT_EXCEL_PATH = r"C:\material_inward\dms_staging\document_links.xlsx"


def main():
    if len(sys.argv) < 3:
        print("Usage: dms_excel_writer.py <file_name> <doc_link> [excel_path]")
        sys.exit(1)

    file_name = sys.argv[1]
    doc_link = sys.argv[2]
    excel_path = sys.argv[3] if len(sys.argv) > 3 else DEFAULT_EXCEL_PATH

    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Document Links"
        ws.append(["File Name", "Document Link"])

    ws.append([file_name, doc_link])
    wb.save(excel_path)
    print(f"Saved: {file_name} -> {doc_link} ({excel_path})")


if __name__ == "__main__":
    main()
