"""
services/dms_links_import.py — imports Contentverse document links from
DMS_LINKS_EXCEL_PATH into the app's own database (dms_document_links table),
so the Documents tab can show/link to the hosted copy.

v16: this closes the loop dms_upload.robot's new "Generate And Save
Document Link" step opened -- that step appends {File Name, Document Link}
rows to an Excel file, but nothing previously read that file back into the
app's database.

Trigger: called directly at the end of services/dms_upload_runner.py's
run_dms_upload(), immediately after a successful upload batch -- NOT on its
own separate schedule. dms_upload_runner.py itself is scheduled (Windows
Task Scheduler, independent of any single Gate In event -- Selenium/
Contentverse automation is too slow/fragile to run synchronously per
record). Chaining the import directly onto that same run means the DB is
never more stale than the upload cadence itself; there is no second cadence
to keep in sync.

Also runnable standalone (python services/dms_links_import.py) for a
one-off catch-up import, e.g. after manually re-running dms_upload.robot.

Safe to re-run: upsert_dms_document_link() is keyed on filename (ON
CONFLICT DO UPDATE), so importing the same Excel file twice is harmless.
"""

import os
import sys
import logging

# Ensure project root is on path when called directly
_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from openpyxl import load_workbook

from config.config import config
from database.connection import init_pool
from database.db_operations import (
    find_history_id_by_consolidated_filename,
    upsert_dms_document_link,
)
from database.migo_operations import get_migo_entry
from database.miro_operations import get_miro_entry
from database.rf_queue_operations import enqueue_rf_job
from config.logger import get_logger

# v20 FIX: this used to call logging.basicConfig() directly with its own
# hardcoded "[dms_links_import]" tag. basicConfig() only ever takes effect
# on the FIRST call in a process and attaches its handler to the ROOT
# logger -- and because this module gets imported early in the worker
# process's chain (rf_queue_worker -> dms_upload_runner -> here), its
# format won that race and got applied to every other module's log lines
# too (they all propagate up to root by default), so e.g. gst_runner/
# gst_operations
# lines were showing up tagged "[dms_links_import]" even though they have
# nothing to do with this file. Switched to the same get_logger() every
# other module in this codebase already uses, which attaches handlers
# directly to each named logger instead of the root logger -- no more
# race, and this script's own logs now also land in the shared rotating
# logs/application.log and logs/errors.log files, which raw basicConfig()
# never wrote to (console only).
logger = get_logger(__name__)


def run_dms_links_import() -> dict:
    """
    Read every row currently in DMS_LINKS_EXCEL_PATH and upsert it into
    dms_document_links. Returns a small summary dict for the caller to log.
    """
    excel_path = config.DMS_LINKS_EXCEL_PATH

    if not os.path.exists(excel_path):
        logger.info(f"No DMS links Excel file at {excel_path} yet — nothing to import")
        return {"imported": 0, "unmatched": 0, "errors": 0}

    try:
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
    except Exception as e:
        logger.error(f"Failed to open {excel_path}: {e}", exc_info=True)
        return {"imported": 0, "unmatched": 0, "errors": 1}

    imported = unmatched = errors = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        file_name, doc_link = (row + (None, None))[:2]
        if not file_name or not doc_link:
            continue

        # dms_upload.robot writes the report name WITHOUT the .pdf
        # extension (os.path.splitext(...)[0]) -- history.consolidated_doc_path
        # is a full path ending in .pdf, so re-add it before matching.
        filename_with_ext = f"{file_name}.pdf" if not str(file_name).lower().endswith(".pdf") else file_name

        try:
            history_id = find_history_id_by_consolidated_filename(filename_with_ext)
            if history_id is None:
                logger.warning(
                    f"Row {row_idx}: no history record found for filename "
                    f"{filename_with_ext!r} — storing link with no history_id"
                )
                unmatched += 1
            ok = upsert_dms_document_link(history_id, filename_with_ext, str(doc_link))
            if ok:
                imported += 1
                if history_id is not None:
                    _catch_up_link_attach_jobs(history_id, str(doc_link))
            else:
                errors += 1
        except Exception as e:
            logger.error(f"Row {row_idx} ({file_name!r}) import failed: {e}", exc_info=True)
            errors += 1

    wb.close()
    logger.info(
        f"DMS links import complete — imported={imported} "
        f"unmatched={unmatched} errors={errors}"
    )
    return {"imported": imported, "unmatched": unmatched, "errors": errors}


def _catch_up_link_attach_jobs(history_id: int, document_link: str) -> None:
    """
    v18: the other half of services/rf_queue_worker.py._enqueue_link_attach.

    That function enqueues migo103_link/migo105_link/miro_link the moment
    MIGO 103 / MIGO 105 / MIRO posts successfully, IF the DMS link already
    exists at that instant -- if it doesn't yet, it marks the step
    'skipped_no_link' instead of enqueueing anything. This function is the
    symmetric other trigger: the moment a link actually lands (here), check
    whether any of those three postings already happened for this
    history_id and are just sitting there waiting on a link -- if so,
    enqueue them now. Same "whichever event happens second" resolution
    already used for pending_po_updates (see
    database/pending_po_operations.py) -- just implemented directly here
    instead of importing services/rf_queue_worker.py, which would create a
    circular import (rf_queue_worker -> dms_upload_runner -> this module).

    Best-effort throughout: never let a problem here fail the link import
    that just succeeded (see caller, wrapped in the row's own try/except).
    """
    migo_entry = get_migo_entry(history_id) or {}
    mat_doc = migo_entry.get("material_doc_number", "")

    if mat_doc and migo_entry.get("migo_103_rf_status") == "success" \
            and migo_entry.get("migo103_link_status") == "skipped_no_link":
        job_id = enqueue_rf_job(
            history_id, "migo103_link",
            {"history_id": history_id, "material_doc_number": mat_doc, "document_link": document_link}
        )
        logger.info(f"Caught up migo103_link for history_id={history_id} (job_id={job_id})")

    if mat_doc and migo_entry.get("migo_105_rf_status") == "success" \
            and migo_entry.get("migo105_link_status") == "skipped_no_link":
        job_id = enqueue_rf_job(
            history_id, "migo105_link",
            {"history_id": history_id, "material_doc_number": mat_doc, "document_link": document_link}
        )
        logger.info(f"Caught up migo105_link for history_id={history_id} (job_id={job_id})")

    miro_entry = get_miro_entry(history_id) or {}
    if mat_doc and miro_entry.get("rf_status") == "success" \
            and miro_entry.get("miro_link_status") == "skipped_no_link":
        job_id = enqueue_rf_job(
            history_id, "miro_link",
            {"history_id": history_id, "material_doc_number": mat_doc, "document_link": document_link}
        )
        logger.info(f"Caught up miro_link for history_id={history_id} (job_id={job_id})")


if __name__ == "__main__":
    init_pool()
    run_dms_links_import()
