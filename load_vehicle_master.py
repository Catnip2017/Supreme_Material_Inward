"""
load_vehicle_master.py — One-time utility to load vehicle master from Excel.

Usage (run from project root):
    python load_vehicle_master.py path\\to\\vehicle_master.xlsx

Expected columns in the Excel file (any order, header row auto-detected):
    Truck Number | Driver Name | Transporter Name | Driver Licence / Driver License

On conflict (same truck+driver combination), existing rows are left unchanged
(ON CONFLICT DO NOTHING), so it's safe to re-run with the same file.

Requires: openpyxl  (pip install openpyxl --break-system-packages)
"""

import sys
import os

# Ensure project root is on path
_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("  pip install openpyxl")
    sys.exit(1)

from database.connection import init_pool
from database.vehicle_master_operations import bulk_insert_vehicle_master

# ── Column name aliases ────────────────────────────────────────────────────────
COLUMN_MAP = {
    "truck number":    "truck_number",
    "truck no":        "truck_number",
    "vehicle number":  "truck_number",
    "vehicle no":      "truck_number",
    "driver name":     "driver_name",
    "driver":          "driver_name",
    "transporter name": "transporter_name",
    "transporter":     "transporter_name",
    "driver licence":  "licence_number",   # British spelling
    "driver license":  "licence_number",   # American spelling
    "licence number":  "licence_number",
    "license number":  "licence_number",
    "licence no":      "licence_number",
    "license no":      "licence_number",
}

REQUIRED_FIELDS = {"truck_number", "driver_name"}


def _normalise_header(raw: str) -> str:
    return raw.strip().lower().replace("_", " ").replace("-", " ")


def _find_header_row(ws) -> tuple[int, dict]:
    """
    Scan the first 10 rows for the header row.
    Returns (row_index_1based, {col_index: field_name}).
    """
    for row_idx in range(1, 11):
        mapping = {}
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            val = str(cell.value or "").strip()
            norm = _normalise_header(val)
            if norm in COLUMN_MAP:
                mapping[col_idx] = COLUMN_MAP[norm]
        if "truck_number" in mapping.values():
            return row_idx, mapping
    raise ValueError(
        "Could not find a header row containing 'Truck Number'. "
        "Check that the column exists and the file is not password-protected."
    )


def load_from_excel(filepath: str) -> int:
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    print(f"Loading: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    header_row, col_map = _find_header_row(ws)
    print(f"  Header row found at row {header_row}")
    print(f"  Mapped columns: {col_map}")

    rows = []
    skipped = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        record = {}
        for col_idx, field_name in col_map.items():
            raw = row[col_idx - 1]  # 0-based tuple
            record[field_name] = str(raw).strip() if raw is not None else ""

        # Skip rows missing required fields
        if not all(record.get(f) for f in REQUIRED_FIELDS):
            skipped += 1
            continue
        rows.append(record)

    wb.close()
    print(f"  Found {len(rows)} data row(s), skipped {skipped} incomplete row(s)")

    if not rows:
        print("Nothing to insert.")
        return 0

    inserted = bulk_insert_vehicle_master(rows)
    print(f"  Inserted {inserted} new row(s) (duplicates skipped).")
    return inserted


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python load_vehicle_master.py <path_to_excel.xlsx>")
        sys.exit(1)

    excel_path = sys.argv[1]
    init_pool()
    total = load_from_excel(excel_path)
    print(f"Done. {total} new vehicle master record(s) added.")